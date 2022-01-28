import re
from lxml import etree
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
'''
st="PARTIAL INTERESTS 2021 CUMULATIVE SUPPLEMENT NO. 1 Copyright &#169; December 2003"
word=re.sub('&[^;]+;', '', st)
sent=word.split(' ')
sent = list(filter(None, sent))
print(sent)
'''
xpath_list=[]
xml_content_list=[]
string_list=[]
def get_xpath(elem,root,temp):
    path = root.getelementpath(elem)
    root_path = '/' + root.getroot().tag
    if path == '.':
        path = root_path
    else:
        path = root_path + '/' + path
    print(path)
    xpath_list.append(path)
    xml_content_list.append(elem.text)
    string_list.append(temp)


def get_xpath_from_xml(temp,sent, count=0):
    with open('VAL_CH_5_Input.xml','r') as out:
        data=out.read()
        out.close()
    root = etree.fromstring(data).getroottree()
    for elem in root.iter():
        if(temp in str(elem.text)):
            print(temp,"  deltatextstring")
            print(elem.text,"  xmlstring")
            count=count+1
    if(count==1):
        get_xpath(elem,root,temp)
        return 1
    else:
        return 0


def prepare_word_for_search(sent):
    print(sent)
    for words in range(len(sent)):
        if(words==0):
            temp=""
            temp=temp+sent[words]
            val=get_xpath_from_xml(temp, sent)
            if(val==1):
                break
        else:
            a=temp+" "+sent[words]
            temp=a
            val=get_xpath_from_xml(temp, sent)
            if(val==1):
                break
    return val

def get_content_from_excel():
    wb = load_workbook('Content.xlsx')
    ws = wb.get_sheet_by_name('Content_issues')
    for row in ws.iter_rows():
        val= [cell.value for cell in row]
        #print(val[0],"gggggggggggggggggggggggggg")
        word = re.sub('&[^;]+;', '', val[0])
        sent = word.split(' ')
        sent = list(filter(None, sent))
        for items in range(len(sent)):
            if (items == 0):
                val = prepare_word_for_search(sent)
                if (val == 1):
                    break
            else:
                prepare_word_for_search(sent[items:])
def writing_into_excel():
    writer = pd.ExcelWriter("Xpath_content.xlsx", engine='xlsxwriter')
    df = pd.DataFrame(
        {'xpath':xpath_list,'xml_content': xml_content_list, 'deltatext': string_list})
    df.to_excel(writer, sheet_name='Style_attributes', index=False)
    writer.save()
get_content_from_excel()
writing_into_excel()
